import csv
import sys
import numpy as np
import openpyxl
import sqlite3
import math
import pandas as pd
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from mpl_toolkits.mplot3d import Axes3D
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
import matplotlib.colors as mcolors


method_descriptions = {
    'MINSUM': {
        'description': (
        'Минимальная сумма (MINSUM)\n'
        '-------------------\n'
        '- Подход: Агрегирует оценки альтернатив на основе суммы их баллов, отдавая предпочтение альтернативам с наименьшим суммарным баллом.\n'
        '- Исходные данные: Требуется матрица решений, состоящая из m альтернатив.\n'
        '- Выходные данные: Одна предпочтительная альтернатива.\n'
        '- Идеально подходит для: Всех задач многокритериальной оптимизации.\n'
        ),
        'args': [
            
        ]
    },
    'MINMAX': {
        'description': (
        'Минимум-Максимум (MINMAX)\n'
        '-----------------------\n'
        '- Подход: Выбирается альтернатива, которая имеет наименьшее максимальное значение, минимизируя наихудший исход.\n'
        '- Исходные данные: Требуется матрица решений, состоящая из m альтернатив.\n'
        '- Выходные данные: Одна предпочтительная альтернатива.\n'
        '- Идеально подходит для: Всех задач многокритериальной оптимизации.\n'
        ),
        'args': [
            
        ]
    },
    'MAXMIN': {
        'description': (
        'Максимум-Минимум (MAXMIN)\n'
        '-----------------------\n'
        '- Подход: Выбирается альтернатива, имеющая наибольшее минимальное значение, максимизирующая наихудший исход.\n'
        '- Исходные данные: Требуется матрица решений, состоящая из m альтернатив.\n'
        '- Выходные данные: Одна предпочтительная альтернатива.\n'
        '- Идеально подходит для: Всех задач многокритериальной оптимизации.\n'
        ),
        'args': [
            
        ]
    },
    'DIP': {
        'description': (
        'Расстояние до исдеальной точки \n'
        '-----------------------\n'
        '- Подход: Для каждого набора критериев вычисляется растояние до точки, в которой отклонения критериев были бы минимальными от текущих.\n'
        '- Исходные данные: Требуется матрица решений, состоящая из m альтернатив .\n'
        '- Выходные данные: Одна предпочтительная альтернатива.\n'
        '- Идеально подходит для: Всех задач многокритериальной оптимизации.\n'
        ),
        'args': [
            
        ]
    }, 'TOPSIS': {
        'description': (
            'Техника упорядочивания предпочтений по сходству с идеальным решением (TOPSIS)\n'
            '-----------------------------------------\n'
            '- Подход: Сравнивает каждую альтернативу с теоретически идеальным решением.\n'
            '- Исходные данные: Требуется матрица оценки, состоящая из m альтернатив и n критериев.\n'
            '- Выходные данные: Ранжирование альтернатив.\n'
            '- Идеально подходит для: Всех задач многокритериальной оптимизации.\n'
        ),
        'args': [
            {
                'name': 'Weights',
                'type': 'multiple',
                'applies_to': 'each_criteria'
            },
        ]
    },
    'WSR': {
        'description': (
            'Weighted Sum Model (WSM) или Weighted Linear Combination\n'
            '----------------------------------------------------\n'
            '- Подход: Присваивает веса в зависимости от важности каждого критерия.\n'
            '- Входы: Баллы каждой альтернативы по каждому критерию.\n'
            '- Выходные данные: Единая общая оценка для каждой альтернативы.\n'
            '- Идеально подходит для: Полно описываемых ситуаций с высокой степенью доступности данных.\n'
        ),
        'args': [
            {
                'name': 'Weights',
                'type': 'multiple',
                'applies_to': 'each_criteria'
            },
            # При необходимости добавьте другие аргументы
        ]
    },
    'ELECTRE': {
        'description': (
            'Метод многокритериальной оценки ELECTRE IV\n'
            '----------------------------------------------------------\n'
            '- Подход: ELECTRE IV представляет собой метод, который использует пороговые значения для определения степени согласованности и несогласованности между альтернативами на основе нормализованных данных.\n'
            '- Исходные данные: Матрица оценок альтернатив по всем критериям, которые нормализуются для учета различных диапазонов значений.\n'
            '- Выходные данные: Матрица альтернатив, отсортированных по степени их предпочтительности на основе согласованности и несогласованности.\n'
            '- Идеально подходит для: Ситуаций, когда требуется оценить альтернативы без явного задания весов критериев, особенно если альтернативы имеют разные типы данных и динамику изменений.\n'
            '- Примечание: Здесь реализован вариант IV поколения метода ELECTRE, не требующий ввода весов критериев и данных о порогах предпочтительности ввиду заранее установленных значений:\n'
            '- q = [0.1] * количество критериев - Порог индифферентности.\n'
            '- p = [0.3] * количество критериев - Порог предпочтения.\n'
            '- v = [0.5] * количество критериев - Порог строгого предпочтения.\n'
            '- Порог согласия = 1 / количество критериев.\n'
            '- Порог несогласия = 0.5.\n'
        ),
        'args': [
            
            ]
    },
    'VIKOR': {
        'description': (
            'VlseKriterijumska Optimizacija I Kompromisno Resenje (VIKOR)\n'
            '----------------------------------------------------------\n'
            '- Подход: Представляет многокритериальный ранжирующий индекс, основанный на конкретной мере "близости" к "идеальному" решению.\n'
            '- Исходные данные: Матрица оценок альтернатив по всем критериям.\n'
            '- Выходные данные: Компромиссное решение и список ранжирования.\n'
            '- Идеально подходит для: Общественные проблемы принятия решений, особенно для решения проблемы распределения водных ресурсов.\n'
        ),
        'args': [
            {
                'name': 'Weights',
                'type': 'multiple',
                'applies_to': 'each_criteria'
            },
            # При необходимости добавьте другие аргументы
        ]
    },
    'ОРП': {
        'description': (
            'Обобщенное решающее правило\n'
            '----------------------------------------------------------\n'
            '- Подход: Представляет собой метод итеративного применения последовательности нормализации таблиц данных и их свертки на основе совместного применения прямых методов максимума, суммы и расстояния до идеальной точки для значений для значений критериев альтернативных решений.\n'
            '- Исходные данные: Матрица данных об альтернативных решениях по всем критериям.\n'
            '- Выходные данные: Матрица альтернатив в порядке их предпочтительности.\n'
            '- Идеально подходит для: задач принятия решений, включающие данные разных типов и направлений динамики.\n'
        ),
        'args': [
            
            ]
    },
    'AHP': {
    'description': (
        'Метод анализа иерархий (AHP)\n'
        '----------------------------------------------------------\n'
        '- Подход: AHP используется для многокритериальной оценки альтернатив, учитывая относительные веса критериев.\n'
        '- Исходные данные: Двумерный массив данных, где каждая строка представляет альтернативу, а столбцы - значения критериев.\n'
        '- Выходные данные: Проранжированный список альтернатив с указанием "Набор №" и взвешенной суммы для каждой альтернативы.\n'
        '- Идеально подходит для: Ситуаций, где важно учитывать относительные значимости критериев при принятии решений.\n'
        '- Примечание: В данной реализации учитываются заданные веса критериев, которые представлены входными данными программы.\n'
    ),
        'args': [
                {
                    'name': 'Weights',
                    'type': 'multiple',
                    'applies_to': 'each_criteria'
                },
            ]
    },
    'CHP': {
    'description': (
        'Метод анализа консенсуса (CHP)\n'
        '----------------------------------------------------------\n'
        '- Подход: CHP используется для многокритериальной оценки, учитывая веса критериев и оценку согласованности решений.\n'
        '- Исходные данные: Двумерный массив данных, представляющий альтернативы и их оценки по критериям, веса критериев заданы входными данными.\n'
        '- Выходные данные: Проранжированный список альтернатив с указанием "Набор №", взвешенной суммы и оценкой согласованности (если применимо).\n'
        '- Идеально подходит для: Ситуаций, где важно не только учитывать веса критериев, но и оценивать согласованность принимаемых решений.\n'
        '- Примечание: В данной реализации рассчитывается оценка согласованности (CR), которая может быть использована для анализа стабильности решений.\n'
    ),
        'args': [
                {
                    'name': 'Weights',
                    'type': 'multiple',
                    'applies_to': 'each_criteria'
                },
            ]
    }
}

class PlotCanvas(FigureCanvas):
    def __init__(self, parent=None):
        self.fig = plt.figure()
        self.ax = self.fig.add_subplot(111, projection='3d')
        super(PlotCanvas, self).__init__(self.fig)
        self.setParent(parent)
        self.plot([])

    def plot(self, data, criterion_index=0, green_indices=None):
        self.ax.clear()
        if len(data) == 0:
            return
        data = np.array(data)
        X = data[:, criterion_index]
        Y = np.prod(np.delete(data, criterion_index, axis=1), axis=1)
        Z = np.prod(data, axis=1)
        
        total_points = len(data)
        if green_indices is None:
            green_indices = []

        colors = [get_color_gradient(i, total_points) for i in range(total_points)]
        for idx in green_indices:
            if idx == green_indices[0]:
                colors[idx] = (0, 1, 0)  # Зеленый цвет для заданных индексов

        # Преобразуем цвета в формат RGBA для использования в scatter
        rgba_colors = np.zeros((len(colors), 4))
        for i, color in enumerate(colors):
            rgba_colors[i, :] = (*color, 1)  # Добавляем альфа-канал, равный 1
        
        self.ax.scatter(X, Y, Z, c=rgba_colors, marker='o')
        self.ax.set_xlabel(f'Критерий {criterion_index + 1}')
        self.ax.set_zlabel('\n\n\nПроизведение значений \n всех критериев')
        self.ax.set_ylabel('\n\n\nПроизведение значений \n остальных критериев')
        self.draw()

    def save_plot(self, filename):
        self.fig.savefig(filename)
        # Сообщение об успешном сохранении
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText("График успешно сохранён!")
        msg.setInformativeText(f"Файл сохранен как: {filename}")
        msg.setWindowTitle("Сохранение графика")
        msg.exec_()

def get_color_gradient(index, total):
    """Возвращает цвет в градиенте от зеленого к красному."""
    normalized_index = index / (total - 1) if total > 1 else 0
    green = np.array(mcolors.to_rgba('green', alpha=1)[:3])
    red = np.array(mcolors.to_rgba('red', alpha=1)[:3])
    return green * (1 - normalized_index) + red * normalized_index



class ResultsWindow(QWidget):
    def __init__(self, data, methods, inputs, parent=None):
        super(ResultsWindow, self).__init__(parent)
        
        self.data = data  # Original data
        self.methods = methods
        self.inputs = inputs
        self.is_pareto = False  # Flag to check whether to show Pareto front or original data
        
        self.initUI()
        
    def initUI(self):
        main_layout = QHBoxLayout()
        left_layout = QVBoxLayout()
        right_layout = QVBoxLayout()

        # 1. Display Data        
        self.data_table = QTableWidget()
        self.data_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.setDataTable()  # Set the table data when the window is loaded
        left_layout.addWidget(self.data_table, 1)
        
        # Create the combo box and labels
        top_layout = QHBoxLayout()
        self.criterion_label = QLabel('Исследуемый критерий:', self)
        self.combo_box = QComboBox(self)
        self.combo_box.addItems([f'# {i+1}' for i in range(len(self.data[0]))])  # Учитываем количество критериев в данных
        self.combo_box.currentIndexChanged.connect(self.update_plot)
        #self.data_table.itemChanged.connect(self.on_data_table_item_changed)
        
        top_layout.addWidget(self.criterion_label)
        top_layout.addWidget(self.combo_box)

        # Add scrollable area for remaining criteria label
        scroll_area = QScrollArea(self)
        scroll_area.setWidgetResizable(True)
        scroll_content = QWidget()
        scroll_layout = QVBoxLayout(scroll_content)
        self.remaining_criteria_label = QLabel('', self)
        scroll_layout.addWidget(self.remaining_criteria_label)
        scroll_area.setWidget(scroll_content)
        right_layout.addLayout(top_layout)
        right_layout.addWidget(scroll_area)

        # Create the plot canvas
        self.canvas = PlotCanvas(self)
        right_layout.addWidget(self.canvas)

        # 2. Results section
        results_layout = QHBoxLayout()

        self.tabs = QTabWidget()
        self.tabs.currentChanged.connect(self.onTabChange)

        # Create and populate results tables for each method
        self.create_results_tables()

        results_layout.addWidget(self.tabs, 1)

        self.pareto_button = QPushButton("Сортировка по Парето")
        self.pareto_button.clicked.connect(self.toggle_pareto)

        self.cancel_button = QPushButton("Отмена", self)
        self.cancel_button.clicked.connect(self.close)

        self.save_button = QPushButton("Сохранить график", self)
        self.save_button.clicked.connect(self.save_plot)
        
        left_layout.addLayout(results_layout, 1)
        left_layout.addWidget(self.pareto_button)
        left_layout.addWidget(self.cancel_button)
        right_layout.addWidget(self.save_button)

        # Add left layout to the main layout
        main_layout.addLayout(left_layout, 1)        

        # Add right layout to the main layout
        main_layout.addLayout(right_layout, 1)

        self.setLayout(main_layout)

        self.update_plot(0)  # Initial update to set labels and plot

    def save_plot(self):
        options = QFileDialog.Options()
        filename, _ = QFileDialog.getSaveFileName(self, "Сохранить график", "", "PNG Files (*.png);;All Files (*)", options=options)
        if filename:
            self.canvas.save_plot(filename)
    
    def onTabChange(self):
        idx = self.tabs.currentIndex()
        cbidx = self.combo_box.currentIndex()
        self.update_plot(cbidx)
    
    def create_results_tables(self):
        # Определяем общее количество критериев, необходимых для всех выбранных методов
        total_criteria = 0
        num_criteria = 0
        criteria_per_method = 0
        criterial_methods = []
        for method in self.methods:
            if method in ["TOPSIS", "WSR", "VIKOR", "AHP", "CHP"]:
                num_criteria = len(self.inputs)                
                total_criteria += 1
                criteria_per_method = int(num_criteria/total_criteria)
                criterial_methods.append(method)
                
        # Создаем вкладки для каждого метода
        for method in self.methods:
            tab = QWidget()
            tab_layout = QHBoxLayout(tab)  # Назначаем макет прямо во вкладке

            results_table = QTableWidget()
            tab_layout.addWidget(results_table)  # Добавляем таблицу в макет
            tab.setLayout(tab_layout)  # Устанавливаем макет для вкладки
            self.tabs.addTab(tab, method)  # Добавляем вкладку в QTabWidget

            # Получаем количество критериев для текущего метода
            if method in ["TOPSIS", "WSR", "VIKOR", "AHP", "CHP"]:
                if method == "TOPSIS":
                    start_index = 0
                    end_index = criteria_per_method
                    filtered_inputs = self.inputs[start_index:end_index]
                    self.update_results(results_table, method, filtered_inputs)    
                if method == "WSR":  
                    start_index = criterial_methods.index("WSR")*criteria_per_method
                    end_index = criterial_methods.index("WSR")*criteria_per_method + criteria_per_method
                    filtered_inputs = self.inputs[start_index:end_index]
                    self.update_results(results_table, method, filtered_inputs)  
                if method == "VIKOR":  
                    start_index = criterial_methods.index("VIKOR")*criteria_per_method
                    end_index = criterial_methods.index("VIKOR")*criteria_per_method + criteria_per_method
                    filtered_inputs = self.inputs[start_index:end_index]
                    self.update_results(results_table, method, filtered_inputs)  
                if method == "AHP":  
                    start_index = criterial_methods.index("AHP")*criteria_per_method
                    end_index = criterial_methods.index("AHP")*criteria_per_method + criteria_per_method
                    filtered_inputs = self.inputs[start_index:end_index]
                    self.update_results(results_table, method, filtered_inputs)  
                if method == "CHP": 
                    start_index = num_criteria - criteria_per_method
                    end_index = num_criteria
                    filtered_inputs = self.inputs[start_index:end_index]
                    self.update_results(results_table, method, filtered_inputs)              
            else:
                self.update_results(results_table, method, None)

    def update_plot(self, index):
        data = self.read_table_data(self.data_table)
        green_indices = self.get_green_indices()  # Получаем зеленые индексы для текущей вкладки
        self.canvas.plot(data, index, green_indices)
        remaining_criteria = [f' {i+1}' for i in range(len(data[0])) if i != index]
        self.remaining_criteria_label.setText(f'Зависимость мультипликативной функции решения от комбинации критериев \n {{ {", ".join(remaining_criteria)} }}')

    def get_green_indices(self):
        # Получаем зеленые индексы для текущей вкладки
        current_tab_index = self.tabs.currentIndex() #if self.tabs.currentIndex() > -1 else 0
        tab = self.tabs.widget(current_tab_index)
        results_table = tab.findChild(QTableWidget)
        if results_table:
            green_indices = []
            for row in range(results_table.rowCount()):
                item = results_table.item(row, 0)
                if item and 'Набор' in item.text():
                    try:
                        index = int(item.text().split('№')[-1]) - 1
                        green_indices.append(index)
                    except ValueError:
                        pass
            return green_indices
        #return []

    def toggle_pareto(self):       
        self.is_pareto = not self.is_pareto  # Toggle the state
        if self.is_pareto:
            self.pareto_button.setText("Данные AS-IS")            
        else:
            self.pareto_button.setText("Сортировка по Парето")
        self.setDataTable()  # Update the table according to the new state
        i = 0
        total_criteria = 0
        num_criteria = 0
        criteria_per_method = 0
        criterial_methods = []
        for method in self.methods:
            if method in ["TOPSIS", "WSR", "VIKOR", "AHP", "CHP"]:
                num_criteria = len(self.inputs)                
                total_criteria += 1
                criteria_per_method = int(num_criteria/total_criteria)
                criterial_methods.append(method)
        
        for method in self.methods:
            tab = self.tabs.widget(i)
            results_table = tab.findChild(QTableWidget)
            if method in ["TOPSIS", "WSR", "VIKOR", "AHP", "CHP"]:
                if method == "TOPSIS":
                    start_index = 0
                    end_index = criteria_per_method
                    filtered_inputs = self.inputs[start_index:end_index]
                    self.update_results(results_table, method, filtered_inputs)    
                if method == "WSR":  
                    start_index = criterial_methods.index("WSR")*criteria_per_method
                    end_index = criterial_methods.index("WSR")*criteria_per_method + criteria_per_method
                    filtered_inputs = self.inputs[start_index:end_index]
                    self.update_results(results_table, method, filtered_inputs)  
                if method == "VIKOR":  
                    start_index = criterial_methods.index("VIKOR")*criteria_per_method
                    end_index = criterial_methods.index("VIKOR")*criteria_per_method + criteria_per_method
                    filtered_inputs = self.inputs[start_index:end_index]
                    self.update_results(results_table, method, filtered_inputs)  
                if method == "AHP":  
                    start_index = criterial_methods.index("AHP")*criteria_per_method
                    end_index = criterial_methods.index("AHP")*criteria_per_method + criteria_per_method
                    filtered_inputs = self.inputs[start_index:end_index]
                    self.update_results(results_table, method, filtered_inputs)  
                if method == "CHP": 
                    start_index = num_criteria - criteria_per_method
                    end_index = num_criteria
                    filtered_inputs = self.inputs[start_index:end_index]
                    self.update_results(results_table, method, filtered_inputs)             
            else:
                self.update_results(results_table, method, None)
            i += 1
        self.update_plot(self.combo_box.currentIndex())

    def setDataTable(self):
        data = self.pareto_dominance(self.data) if self.is_pareto else self.data  # Choose data according to the state
        row_count = len(data)
        col_count = len(data[0])
        self.data_table.setRowCount(row_count)
        self.data_table.setColumnCount(col_count)
        
        for row_index, row_data in enumerate(data):
            for col_index, cell_data in enumerate(row_data):
                self.data_table.setItem(row_index, col_index, QTableWidgetItem(str(round(cell_data, 3))))

    def on_data_table_item_changed(self):
        self.update_plot(self.combo_box.currentIndex())
    
    def read_table_data(self, data_table):
        rows = data_table.rowCount()
        cols = data_table.columnCount()
        data = []
        for row in range(rows):
            row_data = []
            for col in range(cols):
                item = data_table.item(row, col)
                if item and item.text():
                    try:
                        row_data.append(float(item.text()))
                    except ValueError:
                        row_data.append(np.nan)  # Если значение не может быть преобразовано в float, заменяем его на NaN
                else:
                    row_data.append(np.nan)  # Если ячейка пустая, заменяем на NaN
            data.append(row_data)
        return data

    def update_results(self, results_table, method, inputs=None):
        # Создаем словарь названий методов и соответствующих функций
        methods_dict = {
            "MINSUM": self.minsum,
            "MINMAX": self.minmax,
            "MAXMIN": self.maxmin,
            "DIP": self.dip,
            "TOPSIS": self.topsis,
            "WSR": self.wsr,
            "ОРП": self.gsr,
            "ELECTRE": self.electre_iv,
            "VIKOR": self.vikor,
            "AHP": self.ahp,
            "CHP": self.chp
            # Добавьте другие методы по мере необходимости
        }

        # Вызываем соответствующий метод и получаем результаты
        method_name = method.upper()  # Преобразуем название метода в верхний регистр, чтобы соответствовать ключам в methods_dict
        if method_name in methods_dict:
            if method in ["TOPSIS", "WSR", "VIKOR", "AHP", "CHP"]:
                results = methods_dict[method_name](inputs)
            else: results = methods_dict[method_name]()
            results = [[row[0]] + [round(val, 3) if isinstance(val, float) else val for val in row[1:]] for row in results]
        else:
            return
        
        # Теперь заполняем таблицу результатами, полученными из вызова метода
        row_count = len(results)
        col_count = len(results[0]) if results else 0  # Избегаем IndexError в случае пустого списка результатов

        results_table.setRowCount(row_count)
        results_table.setColumnCount(col_count)

        for row_index, row_data in enumerate(results):
            for col_index, col_data in enumerate(row_data):
                if isinstance(col_data, (float, int)):
                    results_table.setItem(row_index, col_index, QTableWidgetItem(str(round(col_data, 3))))
                else:
                    results_table.setItem(row_index, col_index, QTableWidgetItem(str(col_data)))

    def pareto_dominance(self, data):
        if not data or not isinstance(data, list) or len(data) < 2:
            return []
        
        pareto_front = []
        for i, alt1 in enumerate(data):
            dominated = False
            for j, alt2 in enumerate(data):
                if i != j:
                    if all(a >= b for a, b in zip(alt1, alt2)) and any(a > b for a, b in zip(alt1, alt2)):
                        dominated = True
                        break
            if not dominated:
                pareto_front.append(alt1)
        return pareto_front

    def minsum(self):
        data = self.read_table_data(self.data_table)
        minsum_list = []
        for i, row in enumerate(data):
            sum_value = sum(row)
            minsum_list.append([f'Набор №{i+1}'] + [sum_value] + row)
        minsum_list.sort(key=lambda x: x[1])
        return minsum_list

    def minmax(self):
        data = self.read_table_data(self.data_table)
        minmax_list = []
        for i, row in enumerate(data):
            minmax_value = max(row)
            minmax_list.append([f'Набор №{i+1}'] + [minmax_value] + row)
        minmax_list.sort(key=lambda x: x[1])
        return minmax_list

    def maxmin(self):
        data = self.read_table_data(self.data_table)
        maxmin_list = []
        for i, row in enumerate(data):
            maxmin_value = min(row)
            maxmin_list.append([f'Набор №{i+1}'] + [maxmin_value] + row)
        maxmin_list.sort(key=lambda x: x[1], reverse=True)
        return maxmin_list
    
    def dip(self):
        
        data = self.read_table_data(self.data_table)
        num_criteria = len(data[0])
        ideal_point = [min(row[i] for row in data) for i in range(num_criteria)]  # Идеальная точка как минимумы по каждому критерию

        distances = []
        for i, row in enumerate(data):
            dist = math.sqrt(sum((row[j] - ideal_point[j])**2 for j in range(num_criteria)))  # Евклидово расстояние до идеальной точки
            distances.append([f'Набор №{i+1}', dist] + row)

        distances.sort(key=lambda x: x[1])  # Сортировка по возрастанию расстояния до идеальной точки

        return distances
    
    def wsr(self, inputs):
        # Чтение данных из таблицы
        data = self.read_table_data(self.data_table)
        weights = [float(weight.replace(',', '.')) for weight in inputs]  # Заменяем запятые на точки
        

        # Вычисление взвешенных сумм для каждой альтернативы
        weighted_sums = []
        for i, row in enumerate(data):
            weighted_sum = sum(val * weight for val, weight in zip(row, weights))
            weighted_sums.append([f"Набор №{i+1}"] + [weighted_sum] + row)
        
        # Сортировка альтернатив по убыванию взвешенной суммы
        ranked_alternatives = sorted(weighted_sums, key=lambda x: x[1], reverse=True)
        return ranked_alternatives

    def topsis(self, inputs):
        data = self.read_table_data(self.data_table)
        weights = [float(weight.replace(',', '.')) for weight in inputs]  # Заменяем запятые на точки

        normalized_decision_matrix = []
        for col in range(len(data[0])):
            column = [row[col] for row in data]
            norm = (sum(x ** 2 for x in column)) ** 0.5
            normalized_decision_matrix.append([x / norm for x in column])

        weighted_normalized_decision_matrix = [[val * weights[i] for val in col] for i, col in enumerate(normalized_decision_matrix)]
        ideal_solution = [max(col) for col in weighted_normalized_decision_matrix]
        negative_ideal_solution = [min(col) for col in weighted_normalized_decision_matrix]

        separation_measures = []
        for i, row in enumerate(data):
            s_positive = sum((weighted_normalized_decision_matrix[col][i] - ideal_solution[col]) ** 2 for col in range(len(ideal_solution))) ** 0.5
            s_negative = sum((weighted_normalized_decision_matrix[col][i] - negative_ideal_solution[col]) ** 2 for col in range(len(negative_ideal_solution))) ** 0.5
            separation_measures.append([f'Набор №{i+1}'] + [(s_negative / (s_positive + s_negative))] + list(row))

        ranked_alternatives = sorted(separation_measures, key=lambda x: x[1], reverse=True)
        return ranked_alternatives
    
    def gsr(self):
        # Чтение данных из таблицы
        data = self.read_table_data(self.data_table)

        sorted_rows = []
        original_indices = []

        # Создаем список кортежей (индекс, строка)
        indexed_data = [(i, row) for i, row in enumerate(data)]

        # Определение эталонов для каждого столбца
        col_max = [max(col) for col in zip(*data)]
        col_min = [min(col) for col in zip(*data)]
        etalons = [1 / (col_max[i] - col_min[i]) if col_max[i] != col_min[i] else 0 for i in range(len(col_max))]

        # Нормализация данных
        normalized_data = [
            (idx, [x * etalons[j] if etalons[j] != 0 else 0 for j, x in enumerate(row)])
            for idx, row in indexed_data
        ]
        indexed_data = normalized_data

        while len(indexed_data) > 1:          

            valid_rows = []
            for row_index, (idx, row) in enumerate(indexed_data):
                # Проверка условия на отсутствие нескольких единиц в строке
                if row.count(1) < 2:
                    valid_rows.append((idx, row))
                else:
                    # Запись строк, не прошедших условие, в отдельный массив
                    sorted_rows.append(row)
                    original_indices.append(idx)

            # Если нет допустимых строк, завершить цикл
            if not valid_rows:
                break

            results = []
            for idx, row in valid_rows:
                # Вычисление суммы
                row_sum = sum(row)
                
                # Вычисление максимума
                row_max = max(row)
                
                # Вычисление расстояния до идеальной точки
                distance_to_ideal = math.sqrt(sum((x - 1)**2 for x in row))
                
                # Запись значений в результирующий массив
                results.append((idx, [row_sum, row_max, distance_to_ideal]))
            
            transposed_results = [[results[j][1][i] for j in range(len(results))] for i in range(len(results[0][1]))]
            minsum = min(transposed_results[0])
            minmax = min(transposed_results[1])
            mindist = min(transposed_results[2])
            optimized_results = []
            for idx, row in results:
                optimized_results.append((idx, [row[0] / minsum, row[1] / minmax, row[2] / mindist]))
            
            # Обновление данных для следующего цикла
            indexed_data = optimized_results
        
        # Добавление оставшейся строки в отсортированный массив
        if indexed_data:
            sorted_rows.append(indexed_data[0][1])
            original_indices.append(indexed_data[0][0])

        # Добавление информации о номере исходного ряда данных к каждой альтернативе в отсортированном массиве
        sorted_rows_with_indices = []
        for i in range(len(original_indices)):
            sorted_rows_with_indices.append((f"Набор №{original_indices[i]+1}", sorted_rows[i]))
            
        return sorted_rows_with_indices

    def electre_iv(self):
        # Чтение данных из таблицы
        data = self.read_table_data(self.data_table)
        
        num_alternatives = len(data)
        num_criteria = len(data[0])

        # Нормализация данных
        normalized_data = []
        for j in range(num_criteria):
            col = [data[i][j] for i in range(num_alternatives)]
            col_max = max(col)
            normalized_data.append([data[i][j] / col_max for i in range(num_alternatives)])
        
        # Транспонирование нормализованных данных
        normalized_data = list(map(list, zip(*normalized_data)))

        # Определение пороговых значений
        q = [0.1] * num_criteria  # Порог индифферентности
        p = [0.3] * num_criteria  # Порог предпочтения
        v = [0.5] * num_criteria  # Порог строгого предпочтения

        # Создание матриц согласия и несогласия
        concordance_matrix = [[0] * num_alternatives for _ in range(num_alternatives)]
        discordance_matrix = [[0] * num_alternatives for _ in range(num_alternatives)]

        for i in range(num_alternatives):
            for j in range(num_alternatives):
                if i != j:
                    concordance_indices = [k for k in range(num_criteria) if normalized_data[i][k] >= normalized_data[j][k]]
                    discordance_indices = [k for k in range(num_criteria) if normalized_data[i][k] < normalized_data[j][k]]

                    # Согласие
                    concordance_matrix[i][j] = len(concordance_indices) / num_criteria

                    # Несогласие
                    max_discordance = 0
                    for k in discordance_indices:
                        diff = normalized_data[j][k] - normalized_data[i][k]
                        if diff <= q[k]:
                            discordance_value = 0
                        elif q[k] < diff <= p[k]:
                            discordance_value = (diff - q[k]) / (p[k] - q[k])
                        elif p[k] < diff <= v[k]:
                            discordance_value = (diff - p[k]) / (v[k] - p[k])
                        else:
                            discordance_value = 1
                        max_discordance = max(max_discordance, discordance_value)
                    discordance_matrix[i][j] = max_discordance

        # Создание матрицы доминирования
        threshold_concordance = 1 / num_criteria
        threshold_discordance = 0.5

        dominance_matrix = [[0] * num_alternatives for _ in range(num_alternatives)]

        for i in range(num_alternatives):
            for j in range(num_alternatives):
                if i != j:
                    if concordance_matrix[i][j] >= threshold_concordance and discordance_matrix[i][j] <= threshold_discordance:
                        dominance_matrix[i][j] = 1

        # Подсчет баллов доминирования
        dominance_scores = [(f'Альтернатива №{i+1}', sum(dominance_matrix[i]), data[i]) for i in range(num_alternatives)]

        # Сортировка альтернатив по убыванию баллов доминирования
        ranked_alternatives = sorted(dominance_scores, key=lambda x: x[1], reverse=True)

        # Создание информативного результата
        results = []
        for alt, score, alt_data in ranked_alternatives:
            concordant_alternatives = [f"Альтернатива №{j+1}" for j in range(num_alternatives) if dominance_matrix[ranked_alternatives.index((alt, score, alt_data))][j] == 1]
            result_row = [alt,  f"Балл доминирования: {score}"] + list(alt_data) + [f"Доминирует над: {concordant_alternatives}"]
            results.append(result_row)

        return results

    def vikor(self, inputs, q=0.5):
        data = self.read_table_data(self.data_table)
        q_criteria = [float(weight.replace(',', '.')) for weight in inputs]

        # Step 1: Normalize the decision matrix
        normalized_decision_matrix = []
        for col in range(len(data[0])):
            column = [row[col] for row in data]
            min_value = min(column)
            max_value = max(column)
            norm = max_value - min_value
            if norm == 0:
                normalized_column = [0] * len(column)  # Handle the case when all values are the same
            else:
                normalized_column = [(x - min_value) / norm if q_criteria[col] == 1 else (max_value - x) / norm for x in column]
            normalized_decision_matrix.append(normalized_column)

        # Transpose the normalized_decision_matrix to access alternatives easily
        normalized_decision_matrix = list(map(list, zip(*normalized_decision_matrix)))

        # Step 2: Calculate S for each alternative
        S = []
        for i in range(len(data)):
            Si = sum([q_criteria[j] * normalized_decision_matrix[i][j] for j in range(len(data[0]))])
            S.append(Si)

        # Step 3: Calculate R for each alternative
        R = []
        for i in range(len(data)):
            Ri = max([q_criteria[j] * normalized_decision_matrix[i][j] for j in range(len(data[0]))])
            R.append(Ri)

        # Step 4: Calculate VIKOR index for each alternative
        S_min, S_max = min(S), max(S)
        R_min, R_max = min(R), max(R)
        V = []
        for i in range(len(data)):
            Si = (S[i] - S_min) / (S_max - S_min) if S_max != S_min else 0
            Ri = (R[i] - R_min) / (R_max - R_min) if R_max != R_min else 0
            Vi = q * Si + (1 - q) * Ri
            V.append(Vi)

        # Step 5: Rank alternatives based on V
        ranked_alternatives = sorted([(f"Набор №{i+1}", V[i], [f"Критерий {j+1}: {data[i][j]}" for j in range(len(data[0]))]) for i in range(len(data))], key=lambda x: x[1])

        # Step 6: Return ranked alternatives
        return ranked_alternatives
        
    def calculate_pairwise_matrix(self, data, criteria_weights):
        n_criteria = len(criteria_weights)
        pairwise_matrix = np.zeros((n_criteria, n_criteria))
        
        # Calculate pairwise comparisons using criteria weights
        for i in range(n_criteria):
            for j in range(i, n_criteria):
                if i == j:
                    pairwise_matrix[i, j] = 1.0
                else:
                    criteria_i = np.array([row[i] if i < len(row) else 0.0 for row in data])
                    criteria_j = np.array([row[j] if j < len(row) else 0.0 for row in data])
                    criteria_j_nonzero = np.where(criteria_j != 0, criteria_j, 1.0)
                    pairwise_matrix[i, j] = np.sum(criteria_i / criteria_j_nonzero) * criteria_weights[i] / criteria_weights[j]
                    pairwise_matrix[j, i] = 1.0 / pairwise_matrix[i, j]
        
        return pairwise_matrix
    
    def ahp(self, inputs):
        data = self.read_table_data(self.data_table)
        criteria_weights = [float(weight.replace(',', '.')) for weight in inputs]  # Заменяем запятые на точки
    
        pairwise_matrix = self.calculate_pairwise_matrix(data, criteria_weights)
        n = len(pairwise_matrix)
        
        # Step 1: Normalize pairwise comparison matrix
        normalized_matrix = pairwise_matrix / pairwise_matrix.sum(axis=0)
        
        # Step 2: Calculate the weight vector
        weights = normalized_matrix.mean(axis=1)
        
        # Step 3: Normalize the weights
        weights /= weights.sum()
        
        # Calculate weighted sums for each alternative
        weighted_sums = np.zeros(len(data))
        for i in range(len(data)):
            weighted_sums[i] = np.dot(data[i], weights)
        
        # Prepare output with "Набор №" and sort by weighted sums
        ranked_alternatives = []
        for i, row in enumerate(data):
            set_number = f"Набор №{i+1}"
            weighted_sum = weighted_sums[i]
            ranked_alternatives.append([set_number, weighted_sum] + list(row))
        
        ranked_alternatives.sort(key=lambda x: x[1], reverse=True)
        
        return ranked_alternatives

    def chp(self, inputs):
        data = self.read_table_data(self.data_table)
        criteria_weights = [float(weight.replace(',', '.')) for weight in inputs]  # Заменяем запятые на точки
    
        pairwise_matrix = self.calculate_pairwise_matrix(data, criteria_weights)
        n = len(pairwise_matrix)
        
        # Step 1: Calculate row sums and normalize to get initial priority vector
        row_sums = pairwise_matrix.sum(axis=1)
        initial_weights = row_sums / row_sums.sum()
        
        # Step 2: Calculate consistency index (CI)
        ci = (np.dot(pairwise_matrix, initial_weights) - n) / (n - 1)
        
        # Step 3: Calculate consistency ratio (CR)
        random_index = np.array([0, 0.52, 0.89, 1.11, 1.25, 1.35, 1.40, 1.45, 1.49, 1.52])  # Predefined for n = 10
        ri = random_index[n-1]
        cr = ci / ri
        
        # Calculate weighted sums for each alternative
        weighted_sums = np.dot(data, initial_weights)
        
        # Prepare output with "Набор №" and sort by weighted sums
        ranked_alternatives = []
        for i, row in enumerate(data):
            set_number = f"Набор №{i+1}"
            weighted_sum = weighted_sums[i]
            ranked_alternatives.append([set_number, weighted_sum] + list(row))
        
        ranked_alternatives.sort(key=lambda x: x[1], reverse=True)
        
        return ranked_alternatives


class AssessmentWindow(QWidget):
    def __init__(self, data):
        super(AssessmentWindow, self).__init__()
        self.data = data
        self.initUI()

    def initUI(self):
        self.setGeometry(100, 100, 1300, 600)
        self.setWindowTitle('Определение настроек оценки')
        self.selected_methods = []
        self.additional_inputs = {}

        grid_layout = QGridLayout(self)

        # 1. Section for List of Multicriterial Assessment Methods
        self.method_group = QGroupBox("Методы многокритериальной оценки")
        self.method_list = QListWidget()  # Removed self.method_group as parent here
        self.method_list.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)  # Ensure vertical scroll bar appears when needed

        categories_methods = {
            'Прямые методы': ['MINSUM', 'MINMAX', 'MAXMIN', 'DIP', 'ОРП'],
            'Методы взвешенного анализа': ['TOPSIS', 'WSR'],
            'Методы анализа предпочтений': ['ELECTRE'],
            'Методы установления компромиссов': ['VIKOR'],
            'Методы иерархического ранжирования': ['AHP', 'CHP']
        }

        for category, methods in categories_methods.items():
            category_item = QListWidgetItem(category)
            category_item.setFlags(Qt.NoItemFlags)
            self.method_list.addItem(category_item)

            for method in methods:
                method_item = QListWidgetItem()
                checkbox = QCheckBox(method)
                self.method_list.addItem(method_item)
                self.method_list.setItemWidget(method_item, checkbox)
                checkbox.stateChanged.connect(self.display_info)

        # Add method_list to method_group layout
        method_layout = QVBoxLayout(self.method_group)
        method_layout.addWidget(self.method_list)
        self.method_group.setLayout(method_layout)

        grid_layout.addWidget(self.method_group, 0, 0)

        # 2. Multiline Widget for Method Info
        self.info_widget = QTextEdit(self)
        font = self.info_widget.font()
        font.setPointSize(12)
        self.info_widget.setFont(font)
        self.info_widget.setPlaceholderText("Здесь отображается информация о выбранных методах")
        grid_layout.addWidget(self.info_widget, 0, 1)

        # 3. Additional Section for Additional Inputs
        self.additional_inputs_group = QGroupBox("Дополнительные параметры")
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.scroll_content = QWidget()
        self.scroll_layout = QVBoxLayout(self.scroll_content)
        self.form_layout = QFormLayout()
        self.scroll_layout.addLayout(self.form_layout)
        self.scroll_area.setWidget(self.scroll_content)
        self.scroll_area.setFixedHeight(200)  # Set a fixed height for the scroll area
        self.additional_inputs_group.setLayout(QVBoxLayout())
        self.additional_inputs_group.layout().addWidget(self.scroll_area)
        grid_layout.addWidget(self.additional_inputs_group, 1, 0, 1, 2)

        self.resultsButton = QPushButton("Сохранить и Продолжить")
        self.resultsButton.clicked.connect(lambda: self.showResults(self.data, self.selected_methods, self.read_input_fields()))
        grid_layout.addWidget(self.resultsButton, 2, 0, 1, 2)

        grid_layout.setRowStretch(0, 1)  # Stretch the first row so that it takes up the available space
        grid_layout.setRowStretch(1, 0)  # Do not stretch the second row

        self.method_list.itemSelectionChanged.connect(self.display_info)

        
    def showResults(self, data, methods, inputs):
        self.resultsWindow = ResultsWindow(data, methods, inputs)
        self.resultsWindow.show()
        self.center_on_screen()     
    
    def showEvent(self, event):
        self.center_on_screen()
        super().showEvent(event)

    def center_on_screen(self):
        frame_geometry = self.frameGeometry()
        screen_center = QDesktopWidget().availableGeometry().center()
        frame_geometry.moveCenter(screen_center)
        self.move(frame_geometry.topLeft())
    
    def read_input_fields(self):
        input_values = []
        for line_edit in self.additional_inputs.values():
            input_value = line_edit.text()
            input_values.append(input_value)
        return input_values

    def display_info(self):
        self.info_widget.clear()
        # Clear previous widgets
        for i in reversed(range(self.form_layout.count())): 
            widget = self.form_layout.itemAt(i).widget()
            if widget:
                widget.setParent(None)
                
        self.selected_methods = [
            self.method_list.itemWidget(self.method_list.item(i)).text()
            for i in range(self.method_list.count())
            if self.method_list.itemWidget(self.method_list.item(i)) and 
            self.method_list.itemWidget(self.method_list.item(i)).isChecked()
        ]

        if not self.selected_methods:
            self.additional_inputs_group.hide()
            return

        self.additional_inputs.clear()
        
        for method in self.selected_methods:
            self.info_widget.append(method_descriptions[method]['description'])
            if method_descriptions[method]['args']:
                for arg in method_descriptions[method]['args']:
                    self.add_input_field(arg, method)
        
        if self.form_layout.count() > 0:
            self.additional_inputs_group.show()
        else:
            self.additional_inputs_group.hide()

    def add_input_field(self, arg, method):
        name = arg['name']
        arg_type = arg['type']
        applies_to = arg['applies_to']

        if applies_to == 'each_criteria':
            for criterion in range(0, len(self.data[0])):  # Start from 1 to skip the function
                label_str = f"{name} for criterion {criterion} ({method}): "
                line_edit = QLineEdit()
                self.form_layout.addRow(QLabel(label_str), line_edit)
                self.additional_inputs[label_str] = line_edit  # Store the QLineEdit reference
        elif applies_to == 'all':
            if arg_type == 'multiple':
                for criterion in range(0, len(self.data[0])):  # Start from 1 to skip the function
                    label_str = f"{name} for criterion {criterion} ({method}): "
                    line_edit = QLineEdit()
                    self.form_layout.addRow(QLabel(label_str), line_edit)
                    self.additional_inputs[label_str] = line_edit  # Store the QLineEdit reference
            elif arg_type == 'singular':
                label_str = f"{name} ({method}): "
                line_edit = QLineEdit()
                self.form_layout.addRow(QLabel(label_str), line_edit)
                self.additional_inputs[label_str] = line_edit  # Store the QLineEdit reference

class CalculateAlternativesWindow(QDialog):
    def __init__(self):
        super(CalculateAlternativesWindow, self).__init__()

        layout = QVBoxLayout()

        save_button = QPushButton("Сохранить")
        save_button.clicked.connect(self.save_pressed)

        cancel_button = QPushButton("Отмена")
        cancel_button.clicked.connect(self.close)

        layout.addWidget(save_button)
        layout.addWidget(cancel_button)

        self.setLayout(layout)
        self.save_pressed_flag = False

    def save_pressed(self):
        self.save_pressed_flag = True
        self.close()

def open_input_alternatives_window(main_window):
    main_window.hide()
    window = InputAlternativesWindow(main_window)
    window.exec_()
    main_window.show()

class InputAlternativesWindow(QDialog):
    def __init__(self):
        super(InputAlternativesWindow, self).__init__()

        self.data = [['' for _ in range(3)] for _ in range(3)]
        self.initUI()

    def initUI(self):
        self.main_layout = QVBoxLayout(self)
        
        # Setup Buttons
        self.setup_buttons()
        
        # Setup Table
        self.table = QTableWidget(self)
        self.table.setColumnCount(3)
        self.table.setRowCount(3)
        self.table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.table.itemSelectionChanged.connect(self.handle_item_selection)
        
        # Add table to a scroll area
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidget(self.table)
        self.scroll_area.setWidgetResizable(True)
        
        # Place the table and button_layout in a vertical layout
        h_layout = QVBoxLayout()
        h_layout.addLayout(self.button_layout)  # Add button_layout (import and add/remove buttons) to the vertical layout
        h_layout.addWidget(self.scroll_area)  # Add scroll area to the vertical layout
        
        self.main_layout.addLayout(h_layout)  # Add vertical layout to the main layout
        self.main_layout.addLayout(self.save_cancel_layout)  # Add save_cancel_layout below the vertical layout

    def setup_buttons(self):
        # Create Buttons
        self.import_csv_button = QPushButton("Импорт из CSV", self)
        self.import_xlsx_button = QPushButton("Импорт из XLSX", self)
        self.add_remove_column_button = QPushButton("Добавить столбец", self)
        self.add_remove_row_button = QPushButton("Добавить строку", self)

        # Connect Buttons
        self.import_csv_button.clicked.connect(self.import_csv)
        self.import_xlsx_button.clicked.connect(self.import_xlsx)
        self.add_remove_column_button.clicked.connect(self.add_column)
        self.add_remove_row_button.clicked.connect(self.add_row)

        # Button Layout for import and add/remove buttons
        self.button_layout = QVBoxLayout()
        self.button_layout.addWidget(self.import_csv_button)
        self.button_layout.addWidget(self.import_xlsx_button)
        self.button_layout.addWidget(self.add_remove_column_button)
        self.button_layout.addWidget(self.add_remove_row_button)

        # Save and Cancel Buttons
        self.save_button = QPushButton("Сохранить", self)
        self.cancel_button = QPushButton("Отмена", self)
        self.save_button.clicked.connect(self.accept)
        self.cancel_button.clicked.connect(self.reject)

        # Button Layout for Save and Cancel
        self.save_cancel_layout = QHBoxLayout()
        self.save_cancel_layout.addStretch(1)
        self.save_cancel_layout.addWidget(self.save_button)
        self.save_cancel_layout.addWidget(self.cancel_button)
    
    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Escape:
            self.reset_buttons()
            self.table.clearSelection()
        else:
            super(InputAlternativesWindow, self).keyPressEvent(event)
    
    def mousePressEvent(self, event):
        # Detecting if a widget under the cursor is a QPushButton, if not it's assumed to be empty space
        widget = QApplication.widgetAt(QCursor.pos())
        if not isinstance(widget, QPushButton):
            self.reset_buttons()
            self.table.clearSelection()
        super(InputAlternativesWindow, self).mousePressEvent(event)
    
    def reset_buttons(self):
        self.add_remove_row_button.setText("Добавить строку")
        self.add_remove_row_button.disconnect()
        self.add_remove_row_button.clicked.connect(self.add_row)

        self.add_remove_column_button.setText("Добавить столбец")
        self.add_remove_column_button.disconnect()
        self.add_remove_column_button.clicked.connect(self.add_column)
    
    def handle_item_selection(self):
        selected_indices = self.table.selectedIndexes()
        if selected_indices:
            rows = set(index.row() for index in selected_indices)
            columns = set(index.column() for index in selected_indices)
            if len(rows) == 1:
                self.add_remove_row_button.setText("Удалить строку")
                self.add_remove_row_button.disconnect()
                self.add_remove_row_button.clicked.connect(self.delete_row)
            else:
                self.add_remove_row_button.setText("Добавить строку")
                self.add_remove_row_button.disconnect()
                self.add_remove_row_button.clicked.connect(self.add_row)

            if len(columns) == 1:
                self.add_remove_column_button.setText("Удалить столбец")
                self.add_remove_column_button.disconnect()
                self.add_remove_column_button.clicked.connect(self.delete_column)
            else:
                self.add_remove_column_button.setText("Добавить столбец")
                self.add_remove_column_button.disconnect()
                self.add_remove_column_button.clicked.connect(self.add_column)

    def show_error_message(self, message):
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Critical)
        msg_box.setText(message)
        msg_box.setWindowTitle("Error")
        msg_box.exec_()
        
    def import_csv(self):
        options = QFileDialog.Options()
        filePath, _ = QFileDialog.getOpenFileName(self, "Открыть CSV файл", "", "CSV Files (*.csv);;All Files (*)", options=options)
        if filePath:
            try:
                with open(filePath, 'r', newline='', encoding='utf-8') as file:
                    self.data = []
                    reader = csv.reader(file, delimiter=';')
                    for row in reader:
                        if len(row) == 1:
                            row = row[0].split(',')

                        # Filter out non-numeric cells and convert to numbers
                        row = [float(cell.strip()) if cell.strip().replace('.', '', 1).isdigit() else 0 for cell in row]

                        self.data.append(row)

                    self.update_table()
            except FileNotFoundError:
                self.show_error_message("Файл не найден.")
            except UnicodeDecodeError:
                self.show_error_message("Ошибка декодирования файла.")
            except csv.Error as e:
                self.show_error_message(f"Ошибка при чтении CSV файла: {e}")
            except Exception as e:
                self.show_error_message(f"Неожиданная ошибка при чтении CSV файла: {e}")

    def import_xlsx(self):
        options = QFileDialog.Options()
        filePath, _ = QFileDialog.getOpenFileName(self, "Открыть XLSX файл", "", "XLSX Files (*.xlsx);;All Files (*)", options=options)
        if filePath:
            try:
                book = openpyxl.load_workbook(filePath, read_only=True)
                sheet = book.active
                self.data = []

                # Check if the first row is a header
                first_row = next(sheet.iter_rows(values_only=True))
                header_present = all(isinstance(cell, str) for cell in first_row)

                for row in sheet.iter_rows(values_only=True):
                    if header_present:
                        # Skip the first row if it's a header
                        header_present = False
                        continue

                    # Filter out non-numeric cells and convert to numbers
                    processed_row = [float(cell) if cell is not None and str(cell).replace('.', '', 1).isdigit() else 0 for cell in row]
                    self.data.append(processed_row)

                self.update_table()
            except FileNotFoundError:
                self.show_error_message("Файл не найден.")
            except openpyxl.utils.exceptions.InvalidFileException:
                self.show_error_message("Ошибка открытия XLSX файла.")
            except Exception as e:
                self.show_error_message(f"Неожиданная ошибка при чтении XLSX файла: {e}")
    
    def update_table(self):
        self.table.clear()
        row_count = len(self.data)
        col_count = len(self.data[0]) if self.data else 0
        
        self.table.setRowCount(row_count)
        self.table.setColumnCount(col_count)
        
        for row in range(row_count):
            for col in range(col_count):
                item = QTableWidgetItem(str(self.data[row][col]))
                item.setTextAlignment(Qt.AlignCenter)  # Set text alignment to center
                self.table.setItem(row, col, item)

        self.adjust_dialog_size(row_count, col_count)

    def adjust_dialog_size(self, row_count, col_count):
        # Set the maximum number of visible rows and columns before scrolling
        max_visible_rows = 10
        max_visible_cols = 10

        # Calculate table size
        cell_width = 100
        cell_height = 30

        # Determine if we need scrollbars
        table_width = cell_width * min(col_count, max_visible_cols)
        table_height = cell_height * min(row_count, max_visible_rows)

        # Add extra space for scrollbars if necessary
        if col_count > max_visible_cols:
            table_width += 20  # Add space for vertical scrollbar
        if row_count > max_visible_rows:
            table_height += 20  # Add space for horizontal scrollbar

        # Additional space for layouts
        extra_width = 40  # extra space for margins and padding
        extra_height = 60  # extra space for margins and padding

        # Calculate overall size
        total_width = table_width + extra_width 
        total_height = table_height + extra_height + 150

        # Set the dialog size
        self.resize(total_width, total_height)

    def add_column(self):
        for row_data in self.data:
            row_data.append('')
        self.update_table()

    def add_row(self):
        self.data.append(['' for _ in range(self.table.columnCount())])
        self.update_table()

    def delete_column(self):
        col_index = self.table.currentColumn()
        if col_index > -1:
            for row_data in self.data:
                del row_data[col_index]
            self.update_table()

    def delete_row(self):
        row_index = self.table.currentRow()
        if row_index > -1:
            del self.data[row_index]
            self.update_table()
    
    def get_data(self):
        return self.data

class ClickableTextItem(QGraphicsTextItem):
    def __init__(self, text, parent=None, list_widget1=None, list_widget2=None, filename = None):
        super().__init__(text, parent)
        self.setAcceptHoverEvents(True)
        self.filename = filename
        
        # Set Flags
        self.setFlag(QGraphicsItem.ItemIsSelectable)

        # Colors
        self.default_color = QColor(Qt.lightGray)  # Default color
        self.selected_color = QColor(Qt.darkGray)  # Color when selected

        # Initially set to default color
        self.current_color = self.default_color
        self.is_clicked = False  # Initially set to False
        
        # Reference to the list widget
        self.list_widget1 = list_widget1
        self.list_widget2 = list_widget2
        self.setFlag(QGraphicsItem.ItemIsSelectable)
        self.setZValue(11)

    def boundingRect(self):
        return super().boundingRect()  # Should return the correct bounding rect.

    @staticmethod
    def get_related_attributes(attribute_name, table_name, filename):
        connection = sqlite3.connect(filename)
        cursor = connection.cursor()

        # Check if the attribute is a key in the given table
        cursor.execute(f"PRAGMA table_info({table_name})")
        columns = cursor.fetchall()
        is_key = any(col[1] == attribute_name and col[5] for col in columns)  # col[5] is pk column in PRAGMA output

        within_table_attributes = []
        within_database_attributes = []
        if is_key:
            # Within the same table: List all non-key attributes
            within_table_attributes.extend([col[1] for col in columns if not col[5]])  # col[5] is pk column in PRAGMA output

            # Within the database: List all key attributes of other tables that have direct relations to the selected table
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = cursor.fetchall()
            for table in tables:
                table = table[0]
                if table == table_name:  # skip the given table
                    continue
                
                cursor.execute(f"PRAGMA foreign_key_list({table})")
                foreign_keys = cursor.fetchall()
                
                for fk in foreign_keys:
                    ref_table = fk[2]  # Reference table
                    ref_attribute = fk[3]  # Reference attribute (from column)
                                        
                    if ref_table == table_name and ref_attribute == attribute_name:
                        cursor.execute(f"PRAGMA table_info({table})")
                        table_columns = cursor.fetchall()
                        for col in table_columns:
                            is_pk = col[5]  # col[5] is pk column in PRAGMA output
                            col_name = col[1]  # col[1] is the name column in PRAGMA output
                            if is_pk:
                                within_database_attributes.append(col_name)
        else:
            # If not a key, only list all attributes from the same table (both key and non-key)
            within_table_attributes.extend([col[1] for col in columns])

        connection.close()
        """print(is_key, within_table_attributes, within_database_attributes)"""
        return is_key, within_table_attributes, within_database_attributes
    
    def populate_related_attributes(self):
        # Get the reference to the second list widget from the main app or another suitable place.
        second_list_widget = self.list_widget2
        
        second_list_widget = self.list_widget2
        if second_list_widget is not None:
            second_list_widget.clear()
            
            # Retrieve the parent_table object using the parentItem() method.
            parent_table = self.parentItem()
            
            if parent_table:
                attribute_name = self.toPlainText().split(" ")[0]  # Extract attribute name from displayed text
                table_name = parent_table.tableName  # Replace with the actual way to get table name from parent_table object
                
                # Get the related attributes by calling the function
                try:
                    is_key, within_table_attributes, within_database_attributes = ClickableTextItem.get_related_attributes(attribute_name, table_name, self.filename)
                except Exception as e:
                    """print(f"Error in calling get_related_attributes: {str(e)}")"""

                # If the attribute is a key, populate the widget with related attributes
                if is_key:
                    second_list_widget.addItem("В выбранной таблице:")
                    for attribute in within_table_attributes:
                        second_list_widget.addItem(f"  {attribute}")

                    second_list_widget.addItem("В остальной Базе Данных:")
                    for attribute in within_database_attributes:
                        second_list_widget.addItem(f"  {attribute}")
                else:
                    # If not a key, just add the 'Within Table' category
                    second_list_widget.addItem("В выбранной таблице:")
                    for attribute in within_table_attributes:
                        second_list_widget.addItem(f"  {attribute}")


    def mousePressEvent(self, event):
        if self.is_clicked:
            self.is_clicked = False
            self.current_color = self.default_color
            self.remove_from_list()
            self.populate_related_attributes()
        else:
            self.is_clicked = True
            self.current_color = self.selected_color
            self.add_to_list()
            self.populate_related_attributes()

        self.update()  # trigger repaint
        super().mousePressEvent(event)

    def add_to_list(self):
        if self.list_widget1 is None:
            return
        else:
            self.list_widget1.addItem(self.toPlainText())

    def remove_from_list(self):
        if self.list_widget1 is None:
            return
        else:
            items = self.list_widget1.findItems(self.toPlainText(), Qt.MatchExactly)
            for item in items:
                row = self.list_widget1.row(item)
                self.list_widget1.takeItem(row)

    
    def paint(self, painter, option, widget=None):
        # Draw background rectangle
        painter.fillRect(self.boundingRect(), self.current_color)
        super().paint(painter, option, widget)


class RelationGraphicsItem(QGraphicsLineItem):
    def __init__(self, tableItem1, tableItem2, parent=None):
        super().__init__(parent)
        
        self.tableItem1 = tableItem1
        self.tableItem2 = tableItem2
                
        pen = QPen(QBrush(Qt.lightGray), 1)
        self.setPen(pen)

        line = QLineF(tableItem1.sceneBoundingRect().center(), tableItem2.sceneBoundingRect().center())
        self.setLine(line)
        
        tableItem1.relations.append(self)
        tableItem2.relations.append(self)

        self.setZValue(5)  # Lines at the bottom
        self.relations = []
    
class TableGraphicsItem(QGraphicsRectItem):
    def __init__(self, tableName, columns, parent=None, list_widget1=None, list_widget2=None, filename = None):
        super().__init__(parent)
        
        self.tableName = tableName
        self.columns = columns
        self.relations = []
        self.setZValue(9) # Tables just below the titles

        self.setBrush(QBrush(Qt.gray))  # Set to gray or your preferred color
        self.setRect(QRectF(0, 0, 150, 30 + 20 * len(columns)))
        self.setFlag(QGraphicsItem.ItemIsMovable)
        self.setFlag(QGraphicsItem.ItemSendsGeometryChanges)
        
        title_bg = QGraphicsRectItem(0, 0, 150, 30, self)
        title_bg.setBrush(QBrush(Qt.lightGray))
        title_bg.setZValue(10)  # Table titles on top
        
        title = QGraphicsTextItem(tableName, self)
        title.setPos(10, 5)
        title.setZValue(10)  # Table titles on top
        
        for i, (colName, colType) in enumerate(columns):
            textItem = ClickableTextItem(f"{colName} ({colType})", self, list_widget1=list_widget1, list_widget2=list_widget2, filename = filename)
            textItem.setPos(10, 35 + i * 20)
    
    def itemChange(self, change, value):
        if change == QGraphicsItem.ItemPositionHasChanged:  # Notice the change here
            self.update_relations()

        return super().itemChange(change, value)

    def update_relations(self):
        for relation in self.relations:
            rect1 = self.sceneBoundingRect()
            rect2 = relation.tableItem2.sceneBoundingRect() if relation.tableItem1 == self else relation.tableItem1.sceneBoundingRect()

            center1 = rect1.center()
            center2 = rect2.center()

            line = QLineF(center1, center2)

            p1_intersect = self.calculate_intersection(rect1, line)
            p2_intersect = self.calculate_intersection(rect2, line)

            if p1_intersect and p2_intersect:
                new_line = QLineF(p1_intersect, p2_intersect)
                relation.setLine(new_line)
    
    def calculate_intersection(self, rect, line):
        # Generate lines for each edge of the rectangle
        lines = [
            QLineF(rect.topLeft(), rect.topRight()),
            QLineF(rect.topRight(), rect.bottomRight()),
            QLineF(rect.bottomRight(), rect.bottomLeft()),
            QLineF(rect.bottomLeft(), rect.topLeft()),
        ]

        for rect_line in lines:
            intersection_point = QPointF()
            # Find the intersection point between the rectangle edge and the line
            result = line.intersect(rect_line, intersection_point)
            if result == QLineF.BoundedIntersection:
                return intersection_point

        return None

class CalculateAlternativesWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()
    
    def initUI(self):
        # Main Layout
        main_layout = QHBoxLayout(self)
        
        # Left Widget for Graphics View
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        self.scene = QGraphicsScene(self)
        self.view = QGraphicsView(self.scene, self)
        left_layout.addWidget(self.view)
        
        # Right Widget for ListWidgets
        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        
        # First ListWidget and its GroupBox
        self.table_list1 = QListWidget()
        self.table_list1.setStyleSheet("background-color: #F0F0F0;")
        list_group_box1 = QGroupBox("Выбранные аттрибуты:")
        list_group_box1.setLayout(QVBoxLayout())
        list_group_box1.layout().addWidget(self.table_list1)
        right_layout.addWidget(list_group_box1)
        
        # Second ListWidget and its GroupBox
        self.table_list2 = QListWidget()
        self.table_list2.setStyleSheet("background-color: #F0F0F0;")
        list_group_box2 = QGroupBox("Рекомендуемые аттрибуты:")
        list_group_box2.setLayout(QVBoxLayout())
        list_group_box2.layout().addWidget(self.table_list2)
        right_layout.addWidget(list_group_box2)

        #Button for continuing
        continueButton = QPushButton("Далее")
        continueButton.clicked.connect(self.on_continue_clicked)
        right_layout.addWidget(continueButton)
        
        # Add Left and Right Widgets to Main Layout
        main_layout.addWidget(left_widget, 2)  # GraphicsView will take 2 parts of the available space
        main_layout.addWidget(right_widget, 1)  # ListWidgets will take 1 part of the available space
        
        # Window Setup
        self.setGeometry(100, 100, 1300, 600)  # Adjusted the window width
        self.setWindowTitle('Определение аттрибутов из БД')
        
        self.center_on_screen()
        self.show()
        
        # Load Tables
        self.openFileNameDialog()

    def on_continue_clicked(self):
        selected_items = [self.table_list1.item(i).text() for i in range(self.table_list1.count())]
        
        if not selected_items:
            QMessageBox.warning(self, "Ошибка", "Не выбраны атрибуты для продолжения.")
            return
        
        attributes = [item for item in selected_items]
        base_attr = attributes[0].split(" ")[0]  # Get the attribute name from the selected attribute
        
        connection = sqlite3.connect(self.filename)
        cursor = connection.cursor()
    
        tables = connection.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()
        found = False
        for table in tables:
            if found == True: break
            base_table = table[0]
            cursor.execute(f"PRAGMA table_info({base_table})")
            columns = cursor.fetchall()
            for column in columns:
                if column[1] == base_attr:
                    found = True
                    break

        query = f"SELECT {base_table}.{base_attr}"
        
        cursor = connection.cursor()
        join_clauses = []
        found = False
        for attr in attributes[1:]:
            attribute_name = attr.split(" ")[0]  # Get the attribute name from the selected attribute
            tables = connection.execute("SELECT name FROM sqlite_schema WHERE type ='table' AND name NOT LIKE 'sqlite_%'").fetchall()
            for table in tables:
                if found == True: break
                table_name = table[0]
                cursor.execute(f"PRAGMA table_info({table_name})")
                columns = cursor.fetchall()
                for column in columns:
                    col_name = column[1]  # Get the column name
                    if col_name == attribute_name:
                        query += f" , {table_name}.{attribute_name} "
                        join_clauses.append(f"LEFT JOIN {table_name} ON {base_table}.{base_attr} = {table_name}.{base_attr}")
                        found = True
                        break
            
        query += f" FROM {base_table} "
        query += " ".join(join_clauses)
        
        # Execute the query
        try:
            df = pd.read_sql_query(query, connection)
            connection.close()

            # Save the DataFrame to an Excel file
            output_filename = "output.xlsx"
            df.to_excel(output_filename, index=False)
            QMessageBox.information(self, "Успех", f"Данные сохранены в файл {output_filename}")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка SQL", f"Ошибка выполнения запроса SQL: {str(e)}")
            connection.close()

        
    def showEvent(self, event):
        self.center_on_screen()
        super().showEvent(event)

    def center_on_screen(self):
        frame_geometry = self.frameGeometry()
        screen_center = QDesktopWidget().availableGeometry().center()
        frame_geometry.moveCenter(screen_center)
        self.move(frame_geometry.topLeft())
    
    def openFileNameDialog(self):
        options = QFileDialog.Options()
        filename, _ = QFileDialog.getOpenFileName(self, "Открыть файл БД SQLite", "", "SQLite Databases (*.db);;All Files (*)", options=options)
        if filename:
            self.filename = filename
            self.loadTables(filename)
    
    def loadTables(self, dbFile):
        conn = sqlite3.connect(dbFile)
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables = cursor.fetchall()
        
        tableItems = {}
        x = 10
        y = 10
        max_height = 0
        for table in tables:
            tableName = table[0]
            cursor.execute(f"PRAGMA table_info({tableName});")
            columns = [(col[1], col[2]) for col in cursor.fetchall()]
            tableItem = TableGraphicsItem(tableName, columns, list_widget1=self.table_list1, list_widget2=self.table_list2, filename=dbFile)
            tableItem.setPos(x, y)
            self.scene.addItem(tableItem)
            tableItems[tableName] = tableItem
            
            rect = tableItem.rect()
            max_height = max(max_height, rect.height())
            y += max_height + 20
            
            if y > self.view.height() - max_height - 20:
                y = 10
                x += rect.width() + 50
        
        for tableName, tableItem in tableItems.items():
            cursor.execute(f"PRAGMA foreign_key_list({tableName});")
            for relation in cursor.fetchall():
                referencedTable = relation[2]
                if referencedTable in tableItems:
                    lineItem = RelationGraphicsItem(tableItem, tableItems[referencedTable])
                    
                    self.scene.addItem(lineItem)
                    tableItem.relations = getattr(tableItem, 'relations', [])
                    tableItem.relations.append(lineItem)
        conn.close()
        
    def wheelEvent(self, event):
        zoomInFactor = 1.25
        zoomOutFactor = 1 / zoomInFactor
        self.view.setTransformationAnchor(QGraphicsView.NoAnchor)
        self.view.setResizeAnchor(QGraphicsView.NoAnchor)
        oldPos = self.view.mapToScene(event.pos())
        if event.angleDelta().y() > 0:
            zoomFactor = zoomInFactor
        else:
            zoomFactor = zoomOutFactor
        self.view.scale(zoomFactor, zoomFactor)
        newPos = self.view.mapToScene(event.pos())
        delta = newPos - oldPos
        self.view.translate(delta.x(), delta.y())

class App(QMainWindow):
    def __init__(self):
        super(App, self).__init__()

        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.initUI()
        self.dbwindow = None
        self.alternatives_data = None
        self.assessment_window = None

    def initUI(self):
        self.setWindowTitle('Многокритериальная оценка данных')
        layout = QVBoxLayout(self.central_widget)

        calculate_button = QPushButton("Определить альтернативы")
        calculate_button.clicked.connect(self.calculate_alternatives)

        input_button = QPushButton("Ввести альтернативы")
        input_button.clicked.connect(self.input_alternatives)

        self.assessment_button = QPushButton("Определить параметры оценки")
        self.assessment_button.setEnabled(False)
        self.assessment_button.clicked.connect(self.assessment)

        layout.addWidget(calculate_button)
        layout.addWidget(input_button)
        layout.addWidget(self.assessment_button)

        self.setGeometry(100, 100, 400, 80)

        self.center_on_screen()
        self.show()

    def showEvent(self, event):
        self.center_on_screen()
        super().showEvent(event)

    def center_on_screen(self):
        frame_geometry = self.frameGeometry()
        screen_center = QDesktopWidget().availableGeometry().center()
        frame_geometry.moveCenter(screen_center)
        self.move(frame_geometry.topLeft())
    
    def calculate_alternatives(self):
        self.dbwindow = CalculateAlternativesWindow()
        self.dbwindow.show()
        if self.dbwindow.close:
            self.assessment_button.setEnabled(True)

    def input_alternatives(self):
        self.hide()
        window = InputAlternativesWindow()
        if window.exec_() == QDialog.Accepted:  # QDialog.Accepted is returned when 'Save' is pressed
            self.alternatives_data = window.get_data()  # Store table data as state
            self.assessment_button.setEnabled(True)
        self.center_on_screen()
        self.show()

    def assessment(self):
        self.hide()
        self.assessment_window = AssessmentWindow(self.alternatives_data)
        self.assessment_window.show()

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = App()
    sys.exit(app.exec_())
